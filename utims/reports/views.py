from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
from inventory.models import InventoryItem
from tasks.models import Task
from expenses.models import Expense
from datetime import datetime

@login_required
def reports_dashboard(request):
    return render(request, 'reports/dashboard.html')

def generate_inventory_excel(inventory_items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Add headers
    headers = ['Name', 'Category', 'Quantity', 'Price per Unit', 'Last Updated']
    ws.append(headers)
    
    # Add data
    for item in inventory_items:
        ws.append([
            item.name,
            item.category.name,
            item.quantity,
            item.price_per_unit,
            item.last_updated.strftime('%Y-%m-%d')
        ])
    
    return wb

def generate_tasks_excel(tasks):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks Report"
    
    # Add headers
    headers = ['Title', 'Description', 'Status', 'Priority', 'Due Date', 'Assigned To']
    ws.append(headers)
    
    # Add data
    for task in tasks:
        assignee = f"{task.assignee.first_name} {task.assignee.last_name}" if task.assignee else "Unassigned"
        ws.append([
            task.title,
            task.description,
            task.get_status_display(),
            task.get_priority_display(),
            task.due_date.strftime('%Y-%m-%d') if task.due_date else 'No due date',
            assignee
        ])
    
    return wb

def generate_expenses_excel(expenses):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Report"
    
    # Add headers
    headers = ['Title', 'Category', 'Amount', 'Date', 'Description']
    ws.append(headers)
    
    # Add data
    for expense in expenses:
        ws.append([
            expense.title,
            expense.get_category_display(),
            expense.amount,
            expense.date.strftime('%Y-%m-%d'),
            expense.description or '-'
        ])
    
    return wb

def generate_pdf_report(data, headers, title):
    buffer = BytesIO()
    
    # Define margins (in points) - increasing margins to prevent content from going off page
    left_margin = 50
    right_margin = 50
    top_margin = 60
    bottom_margin = 60
    
    # Create document with defined margins
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin
    )
    
    # Calculate available width
    page_width, page_height = letter
    available_width = page_width - (left_margin + right_margin)
    
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph(title, title_style))
    
    # Create table with specific column widths
    table_data = [headers]
    for item in data:
        # Convert each cell to Paragraph for proper text wrapping
        row = []
        for cell in item:
            p = Paragraph(str(cell), styles['Normal'])
            row.append(p)
        table_data.append(row)
    
    # Define column widths based on content - adjust these to change column proportions
    if "Tasks Report" in title:
        # Special column widths for Tasks Report
        col_count = len(headers)
        if col_count == 6:  # Including Assigned To column
            col_widths = [
                available_width * 0.15,  # Title
                available_width * 0.30,  # Description
                available_width * 0.15,  # Status
                available_width * 0.12,  # Priority
                available_width * 0.13,  # Due Date
                available_width * 0.15   # Assigned To
            ]
        else:
            col_widths = [
                available_width * 0.15,  # Title
                available_width * 0.40,  # Description
                available_width * 0.15,  # Status
                available_width * 0.15,  # Priority
                available_width * 0.15   # Due Date
            ]
    elif "Inventory Report" in title:
        # Column widths for Inventory Report
        col_widths = [
            available_width * 0.30,  # Name
            available_width * 0.20,  # Category
            available_width * 0.15,  # Quantity
            available_width * 0.15,  # Price
            available_width * 0.20   # Last Updated
        ]
    else:
        # Column widths for Expenses Report
        col_widths = [
            available_width * 0.20,  # Title
            available_width * 0.20,  # Category
            available_width * 0.15,  # Amount
            available_width * 0.15,  # Date
            available_width * 0.30   # Description
        ]
    
    table = Table(table_data, colWidths=col_widths)
    
    # Add table styles for specific report types
    if "Tasks Report" in title:
        # Use a simpler style for the Tasks Report to match the image
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),  # Left-aligned headers
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white), # White background for data rows
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'), # Left-aligned content
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
    else:
        # Standard styling for other reports
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Headers centered
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertical middle alignment
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf

@login_required
def download_inventory_report(request, format='excel'):
    inventory_items = InventoryItem.objects.all().order_by('name')
    
    if format == 'excel':
        wb = generate_inventory_excel(inventory_items)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=inventory_report.xlsx'
        wb.save(response)
        return response
    
    elif format == 'pdf':
        headers = ['Name', 'Category', 'Quantity', 'Price per Unit', 'Last Updated']
        
        # Format data for PDF display
        data = []
        for item in inventory_items:
            # Format the data, ensuring all values are strings
            data.append([
                str(item.name),
                str(item.category.name),
                str(item.quantity),
                f"${item.price_per_unit}",
                str(item.last_updated.strftime('%Y-%m-%d'))
            ])
            
        pdf = generate_pdf_report(data, headers, "Inventory Report")
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=inventory_report.pdf'
        response.write(pdf)
        return response

@login_required
def download_tasks_report(request, format='excel'):
    tasks = Task.objects.all().order_by('due_date')
    
    if format == 'excel':
        wb = generate_tasks_excel(tasks)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=tasks_report.xlsx'
        wb.save(response)
        return response
    
    elif format == 'pdf':
        headers = ['Title', 'Description', 'Status', 'Priority', 'Due Date', 'Assigned To']
        
        # Truncate long descriptions for PDF display and ensure proper data formatting
        data = []
        for task in tasks:
            # Limit description length to prevent layout issues
            description = task.description or ""
            if len(description) > 100:  # Truncate if too long
                description = description[:97] + "..."
                
            # Get assignee information
            assignee = f"{task.assignee.first_name} {task.assignee.last_name}" if task.assignee else "Unassigned"
                
            # Format the data, ensuring all values are strings
            data.append([
                str(task.title),
                description,
                str(task.get_status_display()),
                str(task.get_priority_display()),
                str(task.due_date.strftime('%Y-%m-%d') if task.due_date else 'No due date'),
                assignee
            ])
            
        pdf = generate_pdf_report(data, headers, "Tasks Report")
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=tasks_report.pdf'
        response.write(pdf)
        return response

@login_required
def download_expenses_report(request, format='excel'):
    expenses = Expense.objects.all().order_by('-date')
    
    if format == 'excel':
        wb = generate_expenses_excel(expenses)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=expenses_report.xlsx'
        wb.save(response)
        return response
    
    elif format == 'pdf':
        headers = ['Title', 'Category', 'Amount', 'Date', 'Description']
        
        # Format data for PDF display
        data = []
        for expense in expenses:
            # Limit description length to prevent layout issues
            description = expense.description or "-"
            if len(description) > 100:  # Truncate if too long
                description = description[:97] + "..."
                
            # Format the data, ensuring all values are strings
            data.append([
                str(expense.title),
                str(expense.get_category_display()),
                f"${expense.amount}",
                str(expense.date.strftime('%Y-%m-%d')),
                description
            ])
            
        pdf = generate_pdf_report(data, headers, "Expenses Report")
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=expenses_report.pdf'
        response.write(pdf)
        return response
